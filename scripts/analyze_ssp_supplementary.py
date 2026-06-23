"""
Supplementary checks on the SSP workbook:
  - Confirm the Images column is empty (images were embedded and skipped).
  - Quantify duplicate model codes (same code, different rows -> variants).
  - Quantify MRP vs SSP coverage and the MRP -> SSP price relationship.
  - Verify there is no tier classification anywhere in the workbook.
  - Detect size/dimension strings inside Description (e.g., 'size: 13.5*16.3*7cm').
"""

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

XLSX_PATH = Path("/home/z/my-project/upload/SSP Final OLD (12).xlsx")
OUT_DIR = Path("/home/z/my-project/scripts/output")

MODEL_CODE_RE = re.compile(r"\b([A-Z]{2,5})\s*[-]?\s*(\d{2,6})\b")
SIZE_RE = re.compile(r"size\s*[:]?\s*([0-9.\s*x×*/,a-z%-]{3,60})", re.IGNORECASE)
TIER_KEYWORDS = ["essential", "premium", "luxury", "economy", "deluxe", "elite"]


def main():
    wb = openpyxl.load_workbook(filename=str(XLSX_PATH), read_only=True, data_only=True, keep_links=False)
    ws = wb["New Laxreee data"]

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx <= 3:
            continue
        cells = [str(v).strip() if v is not None else None for v in list(row)[:8]]
        if not any(cells):
            continue
        rows.append((row_idx, *cells))
    wb.close()

    model_no_idx, desc_idx, disc_idx, discp_idx, mrp_idx, ssp_idx, color_idx, img_idx = 0, 1, 2, 3, 4, 5, 6, 7

    # 1. Images column
    images_nonempty = [r for r in rows if r[img_idx + 1]]  # +1 because first elem is row_idx
    print(f"Images column non-empty cells: {len(images_nonempty)} / {len(rows)}")
    if images_nonempty:
        for r in images_nonempty[:5]:
            print(f"   sample img cell r{r[0]}: {r[img_idx + 1]!r}")

    # 2. Duplicate model codes
    code_to_rows = defaultdict(list)
    for r in rows:
        model = r[model_no_idx + 1]
        if not model:
            continue
        m = MODEL_CODE_RE.search(model)
        if not m:
            continue
        code = f"{m.group(1).upper()}-{m.group(2)}"
        code_to_rows[code].append(r)

    dup_codes = {c: rs for c, rs in code_to_rows.items() if len(rs) > 1}
    print(f"\nDistinct model codes: {len(code_to_rows)}")
    print(f"Model codes appearing >1 time (duplicates/variants): {len(dup_codes)}")
    print("Top 10 duplicate codes (by occurrence count):")
    for code, rs in sorted(dup_codes.items(), key=lambda x: -len(x[1]))[:10]:
        print(f"   {code}: {len(rs)} rows; colors: {[r[color_idx + 1] for r in rs]}")
        for r in rs[:3]:
            print(f"      r{r[0]}: MRP={r[mrp_idx + 1]} SSP={r[ssp_idx + 1]} color={r[color_idx + 1]} desc={(r[desc_idx + 1] or '')[:50]}")

    # 3. MRP vs SSP coverage
    mrp_filled = sum(1 for r in rows if r[mrp_idx + 1] not in (None, ""))
    ssp_filled = sum(1 for r in rows if r[ssp_idx + 1] not in (None, ""))
    print(f"\nMRP populated: {mrp_filled} / {len(rows)}")
    print(f"SSP populated: {ssp_filled} / {len(rows)}")

    # 4. Price relationship SSP = MRP * (1 - discount%/100)
    def to_num(s):
        if s is None or s == "":
            return None
        try:
            return float(s)
        except (ValueError, TypeError):
            return None
    mismatch = 0
    checked = 0
    float_noise = 0
    for r in rows:
        mrp = to_num(r[mrp_idx + 1])
        ssp = to_num(r[ssp_idx + 1])
        discp = to_num(r[discp_idx + 1])
        if mrp and ssp and discp is not None:
            expected = mrp * (1 - discp / 100.0)
            checked += 1
            if abs(expected - ssp) > 1.0:
                mismatch += 1
            elif abs(expected - ssp) > 0.01:
                float_noise += 1
    print(f"\nPrice-formula check: SSP == MRP*(1 - DISCOUNT%/100)")
    print(f"   rows checked: {checked}")
    print(f"   float-noise matches (diff 0.01..1.0): {float_noise}")
    print(f"   mismatches (>1.0): {mismatch}")

    # 5. Size strings in descriptions
    size_hits = 0
    sample_sizes = []
    for r in rows:
        desc = r[desc_idx + 1] or ""
        m = SIZE_RE.search(desc)
        if m:
            size_hits += 1
            if len(sample_sizes) < 10:
                sample_sizes.append(m.group(1).strip())
    print(f"\nDescriptions containing a 'size:' token: {size_hits} / {len(rows)}")
    print("Sample size strings:")
    for s in sample_sizes:
        print(f"   {s!r}")

    # 6. Final tier verification: scan ALL cells (not just description) for tier keywords
    wb2 = openpyxl.load_workbook(filename=str(XLSX_PATH), read_only=True, data_only=True, keep_links=False)
    ws2 = wb2["New Laxreee data"]
    tier_total = 0
    tier_examples = []
    for row_idx, row in enumerate(ws2.iter_rows(values_only=True), start=1):
        for col_idx, v in enumerate(row, start=1):
            if not isinstance(v, str):
                continue
            vl = v.lower()
            for kw in TIER_KEYWORDS:
                if re.search(r"\b" + re.escape(kw) + r"\b", vl):
                    tier_total += 1
                    if len(tier_examples) < 12:
                        tier_examples.append({"row": row_idx, "col": col_idx, "kw": kw, "value": v[:80]})
                    break
    wb2.close()
    print(f"\nFinal tier-keyword scan across ALL cells of the sheet:")
    print(f"   total cells containing a tier keyword: {tier_total}")
    for ex in tier_examples:
        print(f"   r{ex['row']} c{ex['col']} [{ex['kw']}]: {ex['value']!r}")

    out = {
        "total_rows": len(rows),
        "images_nonempty": len(images_nonempty),
        "distinct_model_codes": len(code_to_rows),
        "duplicate_model_codes": len(dup_codes),
        "top_duplicate_codes": [
            {"code": c, "count": len(rs), "colors": [r[color_idx + 1] for r in rs]}
            for c, rs in sorted(dup_codes.items(), key=lambda x: -len(x[1]))[:15]
        ],
        "mrp_populated": mrp_filled,
        "ssp_populated": ssp_filled,
        "price_formula_checked": checked,
        "price_formula_float_noise": float_noise,
        "price_formula_mismatches": mismatch,
        "descriptions_with_size": size_hits,
        "tier_keyword_total_cells": tier_total,
        "tier_keyword_examples": tier_examples,
    }
    with open(OUT_DIR / "ssp_supplementary.json", "w", encoding="utf-8") as f:
        json.dump(out, f, indent=2, ensure_ascii=False, default=str)
    print("\nSupplementary JSON written.")


if __name__ == "__main__":
    main()
