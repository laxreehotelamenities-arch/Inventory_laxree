"""
Analyze LaxRee SSP pricing workbook (SSP Final OLD (12).xlsx).

Goals:
  1. List all sheets with row/col counts.
  2. Sample first 5 rows of each sheet.
  3. Identify how the 3-tier (Essential / Premium / Luxury) classification is represented.
  4. Identify product categories.
  5. Extract sample products with name / model code / tier / price / attributes.
  6. Note English column headers usable as field names.
  7. Find model codes (LR-prefix or similar).
  8. Count total products.

Memory notes:
  - The file is ~50MB and may contain embedded images. We open with read_only=True
    and data_only=True so openpyxl never loads images and only materializes cells
    we actually iterate over.
"""

import json
import re
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

XLSX_PATH = Path("/home/z/my-project/upload/SSP Final OLD (12).xlsx")
OUT_DIR = Path("/home/z/my-project/scripts/output")
OUT_DIR.mkdir(parents=True, exist_ok=True)

TIER_KEYWORDS = ["essential", "premium", "luxury", "economy", "standard", "deluxe", "elite"]

# Match model codes like LR-1234, LR1234, LAX-001, LAXREE-001, etc.
MODEL_CODE_RE = re.compile(r"\b(LR|LAX|LAXREE)[-\s]?(\d{2,6})\b", re.IGNORECASE)
# A looser fallback: any token that looks like CODE-123 / CODE123
LOOSE_CODE_RE = re.compile(r"\b([A-Z]{2,5})[-\s]?(\d{2,6})\b")


def safe_cell(value):
    """Make a cell value JSON-safe and trimmed."""
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    s = str(value).strip()
    return s if s else None


def is_header_like(row_values):
    """A row is header-like if it has multiple non-empty string cells and no numbers."""
    non_empty = [v for v in row_values if v not in (None, "")]
    if len(non_empty) < 2:
        return False
    string_count = sum(1 for v in non_empty if isinstance(v, str))
    number_count = sum(1 for v in non_empty if isinstance(v, (int, float)))
    return string_count >= 2 and number_count == 0


def scan_sheet(ws):
    """Iterate a worksheet in read-only mode and gather structure info."""
    info = {
        "sheet_name": ws.title,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "first_rows": [],
        "header_rows": [],
        "tier_hits": [],  # cells whose value contains a tier keyword
        "tier_in_sheet_name": None,
        "model_codes": [],
        "column_header_candidates": [],
        "numeric_columns": set(),  # 1-based col indexes that mostly contain numbers
        "product_row_count": 0,
        "all_rows_sample": [],  # first 5 rows raw
    }

    # Tier keywords in sheet name?
    name_lower = ws.title.lower()
    for kw in TIER_KEYWORDS:
        if kw in name_lower:
            info["tier_in_sheet_name"] = kw
            break

    # Iterate rows. We cap at a reasonable max for safety, but the file likely has <50k rows per sheet.
    rows_iter = ws.iter_rows(values_only=True)
    col_value_counts = defaultdict(Counter)  # col_index -> Counter(type name)
    header_candidates_per_col = {}  # col_index -> first non-empty cell

    for row_idx, row in enumerate(rows_iter, start=1):
        if row is None:
            continue
        # Trim to max_col to avoid trailing None pollution
        row = list(row)[: ws.max_column] if ws.max_column else list(row)
        cleaned = [safe_cell(v) for v in row]

        if row_idx <= 5:
            info["all_rows_sample"].append({"row": row_idx, "values": cleaned})

        # Detect header rows (usually row 1 or 2)
        if row_idx <= 6 and is_header_like(cleaned):
            info["header_rows"].append({"row": row_idx, "values": cleaned})

        # Tier keyword search in this row
        for col_idx, val in enumerate(cleaned, start=1):
            if not isinstance(val, str):
                continue
            vl = val.lower()
            for kw in TIER_KEYWORDS:
                if kw in vl:
                    info["tier_hits"].append(
                        {"row": row_idx, "col": col_idx, "value": val}
                    )
                    break

            # Model code search
            m = MODEL_CODE_RE.search(val)
            if m:
                info["model_codes"].append(
                    {"row": row_idx, "col": col_idx, "value": val, "code": m.group(0).upper()}
                )

            # Track column header candidates (first non-empty cell per column)
            if col_idx not in header_candidates_per_col and val:
                header_candidates_per_col[col_idx] = val

        # Track per-column types for "numeric column" detection
        for col_idx, val in enumerate(cleaned, start=1):
            if val is None or val == "":
                continue
            if isinstance(val, bool):
                col_value_counts[col_idx]["bool"] += 1
            elif isinstance(val, (int, float)):
                col_value_counts[col_idx]["num"] += 1
            else:
                col_value_counts[col_idx]["str"] += 1

        # Crude product-row heuristic: a row with >=2 non-empty cells, beyond header rows
        non_empty = [v for v in cleaned if v not in (None, "")]
        if len(non_empty) >= 2 and row_idx > 1:
            info["product_row_count"] += 1

    info["column_header_candidates"] = [
        {"col": c, "value": v} for c, v in sorted(header_candidates_per_col.items())
    ]

    # Numeric columns: a column is "numeric" if >50% of its non-empty values are numbers
    for col_idx, counts in col_value_counts.items():
        total = counts["num"] + counts["str"] + counts["bool"]
        if total >= 3 and counts["num"] / total > 0.5:
            info["numeric_columns"].add(col_idx)
    info["numeric_columns"] = sorted(info["numeric_columns"])

    return info


def main():
    print(f"Opening workbook: {XLSX_PATH}")
    print(f"File size: {XLSX_PATH.stat().st_size / (1024*1024):.2f} MB")
    print("-" * 70)

    # read_only=True skips images, data_only=True reads cached values (not formulas)
    wb = openpyxl.load_workbook(
        filename=str(XLSX_PATH),
        read_only=True,
        data_only=True,
        keep_links=False,
    )

    sheet_names = wb.sheetnames
    print(f"Total sheets: {len(sheet_names)}")
    print("Sheet names:")
    for i, name in enumerate(sheet_names, 1):
        print(f"  {i:2d}. {name}")
    print("-" * 70)

    all_sheet_infos = []
    for name in sheet_names:
        ws = wb[name]
        print(f"\n>>> Scanning sheet: {name!r}")
        info = scan_sheet(ws)
        all_sheet_infos.append(info)
        print(
            f"    rows={info['max_row']} cols={info['max_col']} "
            f"product_rows~={info['product_row_count']} "
            f"tier_in_name={info['tier_in_sheet_name']} "
            f"tier_hits={len(info['tier_hits'])} "
            f"model_codes={len(info['model_codes'])}"
        )

    wb.close()

    # ---- Aggregate analysis ----
    print("\n" + "=" * 70)
    print("AGGREGATE ANALYSIS")
    print("=" * 70)

    # 1. Tier representation summary
    print("\n--- Tier classification representation ---")
    tier_in_name_sheets = [i["sheet_name"] for i in all_sheet_infos if i["tier_in_sheet_name"]]
    print(f"Sheets with tier keyword in name: {len(tier_in_name_sheets)}")
    for s in tier_in_name_sheets:
        print(f"   - {s}")

    sheets_with_tier_hits = [i for i in all_sheet_infos if i["tier_hits"]]
    print(f"\nSheets with tier keywords appearing in cells: {len(sheets_with_tier_hits)}")
    for i in sheets_with_tier_hits[:10]:
        tiers_seen = Counter(h["value"].lower() for h in i["tier_hits"])
        print(
            f"   - {i['sheet_name']}: {len(i['tier_hits'])} tier-cell hits, "
            f"sample values: {list(tiers_seen.keys())[:6]}"
        )

    # Where do tier keywords appear? (col distribution)
    tier_col_distribution = Counter()
    for i in all_sheet_infos:
        for h in i["tier_hits"]:
            tier_col_distribution[h["col"]] += 1
    print(f"\nColumn distribution of tier-keyword cell hits (top 10):")
    for col, cnt in tier_col_distribution.most_common(10):
        print(f"   col {col}: {cnt} hits")

    # 2. Model codes summary
    print("\n--- Model codes ---")
    total_codes = sum(len(i["model_codes"]) for i in all_sheet_infos)
    unique_codes = set()
    code_per_sheet = {}
    for i in all_sheet_infos:
        sheet_codes = {m["code"] for m in i["model_codes"]}
        code_per_sheet[i["sheet_name"]] = len(sheet_codes)
        unique_codes.update(sheet_codes)
    print(f"Total model-code cell hits: {total_codes}")
    print(f"Unique model codes: {len(unique_codes)}")
    print("Sample of 20 unique codes:")
    for c in sorted(unique_codes)[:20]:
        print(f"   {c}")

    print("\nModel codes per sheet (top 15):")
    for name, cnt in sorted(code_per_sheet.items(), key=lambda x: -x[1])[:15]:
        print(f"   {name}: {cnt}")

    # 3. Column header candidates aggregated
    print("\n--- Column header candidates (across sheets) ---")
    header_value_counter = Counter()
    for i in all_sheet_infos:
        for h in i["column_header_candidates"]:
            header_value_counter[h["value"]] += 1
    print("Most common first-row cell values (potential field names), top 40:")
    for val, cnt in header_value_counter.most_common(40):
        print(f"   {cnt:4d}  {val!r}")

    # 4. Total product rows
    total_product_rows = sum(i["product_row_count"] for i in all_sheet_infos)
    print(f"\n--- Total estimated product rows (crude, >=2 non-empty cells, row>1): {total_product_rows}")

    # 5. Per-sheet first-5-row samples (compact)
    print("\n--- First 5 rows per sheet (compact) ---")
    for i in all_sheet_infos:
        print(f"\n### Sheet: {i['sheet_name']}  (rows={i['max_row']}, cols={i['max_col']})")
        for r in i["all_rows_sample"]:
            # Truncate long rows for readability
            vals = r["values"]
            compact = [str(v)[:25] if v is not None else "" for v in vals[:12]]
            print(f"   r{r['row']}: {compact}")

    # ---- Save JSON dump for downstream use ----
    # Convert sets to lists for JSON
    for i in all_sheet_infos:
        if isinstance(i["numeric_columns"], set):
            i["numeric_columns"] = sorted(i["numeric_columns"])

    summary = {
        "file": str(XLSX_PATH),
        "file_size_mb": round(XLSX_PATH.stat().st_size / (1024 * 1024), 2),
        "sheet_count": len(all_sheet_infos),
        "sheet_names": sheet_names,
        "sheets": all_sheet_infos,
        "aggregate": {
            "total_product_rows_estimate": total_product_rows,
            "total_model_code_hits": total_codes,
            "unique_model_codes": len(unique_codes),
            "unique_model_code_samples": sorted(unique_codes)[:50],
            "tier_in_sheet_name_count": len(tier_in_name_sheets),
            "sheets_with_tier_in_name": tier_in_name_sheets,
            "sheets_with_tier_in_cells": [i["sheet_name"] for i in sheets_with_tier_hits],
            "tier_col_distribution": dict(tier_col_distribution.most_common(20)),
            "top_header_values": header_value_counter.most_common(60),
        },
    }

    out_path = OUT_DIR / "ssp_analysis.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2, ensure_ascii=False, default=str)
    print(f"\n\nFull JSON analysis written to: {out_path}")


if __name__ == "__main__":
    main()
