#!/usr/bin/env python3
"""
Build the canonical inventory master by merging:
- HTML inventory (Category → Item → Model → Stock)  [PRIMARY structure]
- XLSX images (per model_no)
- PDF/XLSX pricing (per model_no, with tier info as metadata only)

Output: /home/z/my-project/src/data/inventory-master.json
Structure: array of products, each with:
  id, category, item, model_no, colour, balance, inward, dispatched,
  image_url, ssp, mrp, discount_pct, tier (info only), description, source_key

This is the source of truth for the Quick Order cascading dropdown.
"""
import json
import re
import shutil
from pathlib import Path
from collections import defaultdict

ROOT = Path("/home/z/my-project")
HTML = ROOT / "upload/LaxRee_Inventory_2026-06-18 (1) (3).html"
IMAGE_MANIFEST = ROOT / "scripts/output/image_manifest.json"
PDF_ANALYSIS = ROOT / "scripts/output/ssp_pdf_analysis.json"
XLSX_FILE = ROOT / "upload/SSP Final OLD (12).xlsx"
PUBLIC_IMAGES = ROOT / "public/product-images"
PUBLIC_IMAGES.mkdir(parents=True, exist_ok=True)

OUT = ROOT / "src/data/inventory-master.json"

# ---------- Helpers ----------
def norm_model(s: str) -> str:
    """LRWT - 145 → LRWT-145 (uppercase, dash-collapsed)"""
    if not s:
        return ""
    s = str(s).upper().strip()
    m = re.search(r'(LR[A-Z]{2})\s*-?\s*(\d+)', s)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    # Fallback: keep raw text without spaces
    s = re.sub(r'\s+', '-', s.strip())
    s = re.sub(r'[^A-Z0-9\-]', '', s)
    return s

def norm_color(s: str) -> str:
    if not s:
        return ""
    s = str(s).strip()
    placeholders = {"AS PER PICTURE", "AS PER SELECTION", "AS PER CATALOGUE", "AS PER CATALOG"}
    if s.upper() in placeholders:
        return "Multi"
    return s

def parse_size(desc: str) -> str:
    if not desc:
        return ""
    m = re.search(r'(\d+(?:\.\d+)?\s*[*xX×]\s*\d+(?:\.\d+)?(?:\s*[*xX×]\s*\d+(?:\.\d+)?)?\s*(?:mm|cm|m|MM|CM|M)?)', desc)
    if m:
        return m.group(1).strip()
    m = re.search(r'size[:\s]+([^.)]+)', desc, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return ""

# ---------- 1. Load HTML inventory ----------
print("1. Loading HTML inventory...")
text = HTML.read_text(encoding='utf-8', errors='ignore')
m = re.search(r'EMBEDDED_DATA\s*=\s*(\{.*?\});\s*\n', text, re.DOTALL)
data = json.loads(m.group(1))
smap = data.get("smap", {})
print(f"   Loaded {len(smap)} inventory records")

# ---------- 2. Load image manifest ----------
print("2. Loading image manifest...")
with open(IMAGE_MANIFEST) as f:
    img_data = json.load(f)
images_by_model = defaultdict(list)
for entry in img_data["mappings"]:
    mn = norm_model(entry.get("model_no_clean", "").replace("_", "-"))
    if mn:
        images_by_model[mn].append(entry)
print(f"   Loaded {len(img_data['mappings'])} image mappings")

# ---------- 3. Load PDF pricing (with tier info) ----------
print("3. Loading PDF pricing data...")
with open(PDF_ANALYSIS) as f:
    pdf_data = json.load(f)
pdf_pricing = {}  # model_norm -> best pricing record (with tier)
for p in pdf_data["all_products"]:
    mn = p.get("model_no_norm") or norm_model(p.get("model_no", ""))
    if not mn:
        continue
    ssp = p.get("ssp")
    if ssp is not None:
        try:
            ssp = round(float(ssp), 2)
        except (ValueError, TypeError):
            ssp = None
    rec = {
        "ssp": ssp,
        "tier": p.get("tier_filled") or p.get("tier") or "Standard",
        "category_pdf": p.get("product_filled") or p.get("product", ""),
        "description_pdf": p.get("description", ""),
        "color_pdf": norm_color(p.get("color", "")),
    }
    # If we already have a record, prefer the one with tier info
    if mn in pdf_pricing:
        existing = pdf_pricing[mn]
        if rec["tier"] != "Standard" and existing["tier"] == "Standard":
            pdf_pricing[mn] = rec
        elif rec["ssp"] and not existing["ssp"]:
            pdf_pricing[mn] = rec
    else:
        pdf_pricing[mn] = rec
print(f"   Loaded pricing for {len(pdf_pricing)} models")

# ---------- 4. Load XLSX pricing (MRP, discount) ----------
print("4. Loading XLSX pricing (MRP/discount)...")
xlsx_pricing = {}
try:
    from openpyxl import load_workbook
    wb = load_workbook(XLSX_FILE, read_only=True, keep_links=False, data_only=True)
    ws = wb["New Laxreee data"]
    for row in ws.iter_rows(min_row=4, max_row=900, values_only=True):
        if not row or len(row) < 7:
            continue
        model_no = str(row[0] or "").strip()
        if not model_no:
            continue
        mn = norm_model(model_no)
        if not mn:
            continue
        # Pick first row we find for this model with valid MRP/discount
        if mn in xlsx_pricing:
            continue
        try:
            mrp = round(float(row[4])) if row[4] else None
        except (ValueError, TypeError):
            mrp = None
        try:
            ssp = round(float(row[5])) if row[5] else None
        except (ValueError, TypeError):
            ssp = None
        try:
            discount = round(float(row[3]), 2) if row[3] else None
        except (ValueError, TypeError):
            discount = None
        xlsx_pricing[mn] = {
            "mrp": mrp,
            "ssp_xlsx": ssp,
            "discount_pct": discount,
            "description_xlsx": str(row[1] or "").strip(),
        }
    wb.close()
    print(f"   Loaded XLSX pricing for {len(xlsx_pricing)} models")
except Exception as e:
    print(f"   WARNING: XLSX load failed: {e}")

# ---------- 5. Build unified products ----------
print("\n5. Building unified products...")

products = []
for key, rec in smap.items():
    category = rec.get("category", "").strip()
    item = rec.get("item", "").strip()
    model = rec.get("model", "").strip()
    colour = rec.get("colour", "").strip()
    balance = rec.get("balance", 0)
    inward = rec.get("inward", 0)
    dispatched = rec.get("dispatched", 0)

    if not category or not item:
        continue

    model_norm = norm_model(model)
    color_norm = norm_color(colour)

    # Match image (try model + colour match first, fall back to model only)
    image_url = None
    candidates = images_by_model.get(model_norm, [])
    if candidates:
        # Prefer ones that match colour; else pick first
        best = None
        for c in candidates:
            cm = c.get("model_no_clean", "").upper()
            if color_norm and color_norm.upper() in cm:
                best = c
                break
        if not best:
            # Sort: prefer non-duplicate, smallest row
            candidates.sort(key=lambda x: (x.get("image_index_in_row", 99), x.get("row", 9999)))
            best = candidates[0]
        src = Path(best["image_path"])
        if src.exists():
            dest = PUBLIC_IMAGES / src.name
            if not dest.exists():
                try:
                    shutil.copy2(src, dest)
                except Exception:
                    pass
            image_url = f"/product-images/{src.name}"

    # Match pricing
    pdf_rec = pdf_pricing.get(model_norm, {})
    xlsx_rec = xlsx_pricing.get(model_norm, {})

    ssp = pdf_rec.get("ssp") or xlsx_rec.get("ssp_xlsx")
    if ssp is not None:
        try:
            ssp = round(float(ssp), 2)
        except (ValueError, TypeError):
            ssp = None
    mrp = xlsx_rec.get("mrp")
    discount_pct = xlsx_rec.get("discount_pct")

    # Tier (informational — from PDF, used for alternative suggestions)
    tier = pdf_rec.get("tier") or "Standard"

    # Description (prefer PDF, fallback to XLSX)
    description = pdf_rec.get("description_pdf") or xlsx_rec.get("description_xlsx") or item

    # Unique id
    sku_key = f"{model}__{colour}".replace(" ", "_").upper()
    pid = f"inv-{len(products)+1}-{sku_key[:40]}"

    products.append({
        "id": pid,
        "category": category,
        "item": item,
        "model_no": model,
        "model_norm": model_norm,
        "colour": colour,
        "color": color_norm,
        "balance": balance,
        "inward": inward,
        "dispatched": dispatched,
        "stock_qty": balance,
        "in_stock": balance > 0,
        "tier": tier,
        "ssp": ssp,
        "mrp": mrp,
        "discount_pct": discount_pct,
        "description": description,
        "size": parse_size(description),
        "image_url": image_url,
        "name": item,  # Use item name as primary name
        "source_key": key,
    })

# Sort products by category → item → model_no → colour
products.sort(key=lambda p: (p["category"], p["item"], p["model_no"], p["colour"]))

print(f"   Total products: {len(products)}")
print(f"   With image: {sum(1 for p in products if p['image_url'])}")
print(f"   With price (ssp): {sum(1 for p in products if p['ssp'])}")
print(f"   In stock (>0): {sum(1 for p in products if p['stock_qty'] > 0)}")
print(f"   Out of stock: {sum(1 for p in products if p['stock_qty'] == 0)}")

# Verify hierarchy
cats = sorted(set(p["category"] for p in products))
print(f"\n   Categories ({len(cats)}): {cats}")
print(f"   Total unique items: {len(set((p['category'], p['item']) for p in products))}")
print(f"   Total unique models: {len(set(p['model_norm'] for p in products))}")

# ---------- 6. Save ----------
OUT.parent.mkdir(parents=True, exist_ok=True)
with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(products, f, ensure_ascii=False, indent=2)
print(f"\n6. Saved to: {OUT}")
print(f"   Size: {OUT.stat().st_size / 1024:.1f} KB")

# Save a hierarchy summary for quick reference
hierarchy_summary = {
    "categories": {},
    "total_categories": len(cats),
    "total_items": len(set((p["category"], p["item"]) for p in products)),
    "total_skus": len(products),
}
for cat in cats:
    items_in_cat = sorted(set(p["item"] for p in products if p["category"] == cat))
    hierarchy_summary["categories"][cat] = {
        "item_count": len(items_in_cat),
        "items": items_in_cat,
    }
summary_path = ROOT / "src/data/inventory-hierarchy.json"
with open(summary_path, 'w', encoding='utf-8') as f:
    json.dump(hierarchy_summary, f, ensure_ascii=False, indent=2)
print(f"   Hierarchy summary: {summary_path}")

print("\nDONE.")
