#!/usr/bin/env python3
"""
Extract embedded product images from an XLSX file and map them to product rows.

Strategy:
  - Treat the .xlsx as a zip archive (read-only, never modify the original).
  - Parse xl/drawings/drawing1.xml to read every <xdr:twoCellAnchor> /
    <xdr:oneCellAnchor> element, extracting the anchor row (0-indexed) and the
    embedded image relationship id (r:embed="rIdN").
  - Parse xl/drawings/_rels/drawing1.xml.rels to map rId -> media file path
    inside xl/media/.
  - Use openpyxl (read_only=True, keep_links=False) to read column 1 (Model No)
    for every product row (1-indexed rows 4..900).
  - For every unique (row, rId) anchor, read the image bytes from the zip,
    compute a sha256 hash so we can identify identical source images shared
    across multiple rows, resize to max 1200 px on the long edge, and save
    under /home/z/my-project/download/product-images/ with a clean filename.
  - Emit a JSON manifest at /home/z/my-project/scripts/output/image_manifest.json.

Memory notes:
  - The xlsx is ~50 MB, mostly images. We never load the whole archive into
    memory: zipfile lets us stream individual entries, and we process one
    image at a time.
  - openpyxl read_only mode streams rows lazily.

Usage:
    python3 extract_images.py
"""

from __future__ import annotations

import hashlib
import io
import json
import os
import re
import sys
import zipfile
from collections import defaultdict
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from PIL import Image

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
XLSX_PATH = "/home/z/my-project/upload/SSP Final OLD (12).xlsx"
OUTPUT_DIR = "/home/z/my-project/download/product-images/"
MANIFEST_PATH = "/home/z/my-project/scripts/output/image_manifest.json"

SHEET_NAME = "New Laxreee data"
PRODUCT_FIRST_ROW = 4  # 1-indexed
PRODUCT_LAST_ROW = 900  # 1-indexed
MODEL_NO_COLUMN = 1  # 1-indexed (column A)
IMAGES_COLUMN_0IDX = 7  # 0-indexed -> column 8 ("Images")

MAX_LONG_EDGE = 1200  # px; resize images larger than this on the long edge
JPEG_QUALITY = 88

# Regexes for parsing the drawing XML
ANCHOR_SPLIT_RE = re.compile(r"<xdr:(twoCellAnchor|oneCellAnchor)>")
FROM_BLOCK_RE = re.compile(r"<xdr:from>(.*?)</xdr:from>", re.DOTALL)
ROW_RE = re.compile(r"<xdr:row>(\d+)</xdr:row>")
COL_RE = re.compile(r"<xdr:col>(\d+)</xdr:col>")
EMBED_RE = re.compile(r'r:embed="(rId\d+)"')
REL_RE = re.compile(
    r'<Relationship\s+Id="(rId\d+)"[^>]*Target="([^"]+)"[^>]*>'
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def log(msg: str) -> None:
    print(msg, flush=True)


def parse_rels(rels_xml: str) -> Dict[str, str]:
    """Return {rId: target_path_relative_to_xl/drawings/}."""
    rel_map: Dict[str, str] = {}
    for m in REL_RE.finditer(rels_xml):
        rel_map[m.group(1)] = m.group(2)
    return rel_map


def resolve_media_path(rel_target: str) -> str:
    """Resolve a rels Target (e.g. '../media/image1.png') to an archive path."""
    # rels targets are relative to xl/drawings/, so ../media/X -> xl/media/X
    return os.path.normpath(os.path.join("xl", "drawings", rel_target)).replace(
        os.sep, "/"
    )


def parse_drawing(drawing_xml: str) -> List[dict]:
    """Return list of anchor dicts: {row, col, rId, type}."""
    parts = ANCHOR_SPLIT_RE.split(drawing_xml)
    anchors: List[dict] = []
    # parts[0] is the preamble; then pairs of (anchor_type, body)
    for i in range(1, len(parts), 2):
        anchor_type = parts[i]
        body = parts[i + 1].split("</xdr:twoCellAnchor>")[0].split(
            "</xdr:oneCellAnchor>"
        )[0]
        from_match = FROM_BLOCK_RE.search(body)
        from_inner = from_match.group(1) if from_match else ""
        row_m = ROW_RE.search(from_inner)
        col_m = COL_RE.search(from_inner)
        embed_m = EMBED_RE.search(body)
        anchors.append(
            {
                "type": anchor_type,
                "row": int(row_m.group(1)) if row_m else None,
                "col": int(col_m.group(1)) if col_m else None,
                "rId": embed_m.group(1) if embed_m else None,
            }
        )
    return anchors


def sanitize_model_no(raw: Optional[str]) -> str:
    """
    Clean a model-no string for use as a filename component.
    e.g. 'LRWA - 399' -> 'LRWA-399'
         'LRLI - 467 (WET FLOOR SIGNAGE)' -> 'LRLI-467_WET-FLOOR-SIGNAGE'
    """
    if raw is None:
        return "unknown"
    s = str(raw).strip()
    if not s:
        return "unknown"
    # Collapse 'LRWA - 399' / 'LRWA  -  399' -> 'LRWA-399'
    s = re.sub(r"\s*-\s*", "-", s)
    # Replace parens content separators nicely: 'Foo (Bar)' -> 'Foo_Bar'
    s = re.sub(r"\s*\(", "_", s)
    s = re.sub(r"\)", "", s)
    # Replace whitespace runs with single underscore
    s = re.sub(r"\s+", "_", s)
    # Strip characters that are unsafe on common filesystems
    s = re.sub(r'[\\/:*?"<>|]', "-", s)
    # Collapse repeated underscores / dashes
    s = re.sub(r"_+", "_", s)
    s = re.sub(r"-+", "-", s)
    s = s.strip("_-")
    return s or "unknown"


def read_model_numbers(xlsx_path: str) -> Dict[int, str]:
    """
    Return {1-indexed-row: raw-model-no} for rows 1..PRODUCT_LAST_ROW.
    Uses openpyxl read_only mode to stream rows lazily.
    """
    log(f"[openpyxl] loading workbook read-only: {xlsx_path}")
    wb = load_workbook(xlsx_path, read_only=True, keep_links=False, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}"
        )
    ws = wb[SHEET_NAME]
    log(f"[openpyxl] sheet '{SHEET_NAME}': max_row={ws.max_row}, "
        f"max_col={ws.max_column}")

    model_by_row: Dict[int, str] = {}
    for row_idx, row in enumerate(
        ws.iter_rows(
            min_row=1,
            max_row=PRODUCT_LAST_ROW,
            min_col=MODEL_NO_COLUMN,
            max_col=MODEL_NO_COLUMN,
            values_only=True,
        ),
        start=1,
    ):
        if row and row[0] is not None and str(row[0]).strip():
            model_by_row[row_idx] = str(row[0]).strip()
    wb.close()
    log(f"[openpyxl] read {len(model_by_row)} non-empty model-no cells "
        f"(rows 1..{PRODUCT_LAST_ROW})")
    return model_by_row


def resize_image_bytes(
    data: bytes, src_ext: str
) -> Tuple[bytes, str, int, int, int, int]:
    """
    Open image bytes with PIL, resize if larger than MAX_LONG_EDGE.
    Returns (output_bytes, output_ext, src_w, src_h, out_w, out_h).
    Keeps PNG as PNG (preserves transparency); JPEG/PNG both resize cleanly.
    """
    src_ext_lower = (src_ext or "").lower()
    with Image.open(io.BytesIO(data)) as img:
        src_w, src_h = img.size
        # Convert mode if needed for saving
        out_w, out_h = src_w, src_h
        needs_resize = max(src_w, src_h) > MAX_LONG_EDGE
        if needs_resize:
            ratio = MAX_LONG_EDGE / max(src_w, src_h)
            out_w = max(1, int(round(src_w * ratio)))
            out_h = max(1, int(round(src_h * ratio)))

        # Choose output extension based on source format.
        fmt = (img.format or "").upper()
        if fmt == "PNG":
            out_ext = "png"
            # Ensure RGBA / P modes are handled; PIL saves PNG fine.
            save_img = img
            if needs_resize:
                save_img = img.resize((out_w, out_h), Image.LANCZOS)
            buf = io.BytesIO()
            save_img.save(buf, format="PNG", optimize=True)
        elif fmt in ("JPEG", "JPG"):
            out_ext = "jpg"
            save_img = img
            if needs_resize:
                save_img = img.resize((out_w, out_h), Image.LANCZOS)
            # Flatten if necessary (JPEG has no alpha)
            if save_img.mode in ("RGBA", "LA", "P"):
                bg = Image.new("RGB", save_img.size, (255, 255, 255))
                if save_img.mode == "P":
                    save_img = save_img.convert("RGBA")
                bg.paste(save_img, mask=save_img.split()[-1] if save_img.mode in ("RGBA", "LA") else None)
                save_img = bg
            buf = io.BytesIO()
            save_img.save(buf, format="JPEG", quality=JPEG_QUALITY, optimize=True)
        else:
            # Fallback: save as PNG
            out_ext = "png"
            save_img = img
            if needs_resize:
                save_img = img.resize((out_w, out_h), Image.LANCZOS)
            buf = io.BytesIO()
            save_img.save(buf, format="PNG", optimize=True)
        return buf.getvalue(), out_ext, src_w, src_h, out_w, out_h


def file_extension_for(archive_path: str) -> str:
    return archive_path.rsplit(".", 1)[-1].lower() if "." in archive_path else "bin"


# ---------------------------------------------------------------------------
# Main extraction pipeline
# ---------------------------------------------------------------------------
def main() -> int:
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(os.path.dirname(MANIFEST_PATH), exist_ok=True)

    if not os.path.isfile(XLSX_PATH):
        log(f"ERROR: xlsx not found: {XLSX_PATH}")
        return 2

    log("=" * 70)
    log(f"Source xlsx : {XLSX_PATH}")
    log(f"Output dir  : {OUTPUT_DIR}")
    log(f"Manifest    : {MANIFEST_PATH}")
    log("=" * 70)

    # 1) Read model numbers via openpyxl (streaming, read-only)
    model_by_row = read_model_numbers(XLSX_PATH)

    # 2) Open xlsx as zip and parse drawing + rels
    with zipfile.ZipFile(XLSX_PATH, "r") as zf:
        names = zf.namelist()
        media_entries = [n for n in names if n.startswith("xl/media/")]
        log(f"[zip] total archive entries: {len(names)}")
        log(f"[zip] media entries       : {len(media_entries)}")

        # ext & size breakdown
        ext_counts: Dict[str, int] = defaultdict(int)
        total_media_bytes = 0
        for n in media_entries:
            ext_counts[file_extension_for(n)] += 1
            total_media_bytes += zf.getinfo(n).file_size
        log(f"[zip] media ext counts    : {dict(ext_counts)}")
        log(f"[zip] total media bytes   : {total_media_bytes:,} "
            f"({total_media_bytes / 1024 / 1024:.1f} MB)")

        # Parse rels
        with zf.open("xl/drawings/_rels/drawing1.xml.rels") as f:
            rels_xml = f.read().decode("utf-8", errors="replace")
        rel_map = parse_rels(rels_xml)
        log(f"[drawing] relationships   : {len(rel_map)}")

        # Parse drawing
        with zf.open("xl/drawings/drawing1.xml") as f:
            # 1 MB file; safe to read whole
            drawing_xml = f.read().decode("utf-8", errors="replace")
        anchors = parse_drawing(drawing_xml)
        log(f"[drawing] total anchors   : {len(anchors)}")

        # 3) Filter anchors to valid product-image anchors:
        #    - has rId
        #    - rId is in rel_map
        #    - col == IMAGES_COLUMN_0IDX (column 8 = Images)
        #    - row corresponds to a product row (1-indexed 4..900)
        valid_anchors: List[dict] = []
        skipped_no_rid = 0
        skipped_broken_rid = 0
        skipped_wrong_col = 0
        skipped_out_of_range = 0
        for a in anchors:
            if not a["rId"]:
                skipped_no_rid += 1
                continue
            if a["rId"] not in rel_map:
                skipped_broken_rid += 1
                continue
            if a["col"] != IMAGES_COLUMN_0IDX:
                skipped_wrong_col += 1
                continue
            row_1idx = a["row"] + 1
            if row_1idx < PRODUCT_FIRST_ROW or row_1idx > PRODUCT_LAST_ROW:
                skipped_out_of_range += 1
                continue
            valid_anchors.append(a)
        log(f"[filter] valid product-image anchors: {len(valid_anchors)}")
        log(f"[filter] skipped (no rId)       : {skipped_no_rid}")
        log(f"[filter] skipped (broken rId)   : {skipped_broken_rid}")
        log(f"[filter] skipped (wrong col)    : {skipped_wrong_col}")
        log(f"[filter] skipped (out of range) : {skipped_out_of_range}")

        # 4) Dedupe per (row, rId). Multiple identical anchors on the same row
        #    (same rId) collapse to a single output file. Multiple distinct
        #    rIds on the same row each get their own file with a suffix.
        row_to_rids: Dict[int, List[str]] = defaultdict(list)
        seen_pairs: set = set()
        for a in valid_anchors:
            key = (a["row"], a["rId"])
            if key in seen_pairs:
                continue
            seen_pairs.add(key)
            row_to_rids[a["row"]].append(a["rId"])

        # Sort rIds within each row deterministically (by numeric rId value)
        def rid_num(rid: str) -> int:
            m = re.match(r"rId(\d+)$", rid)
            return int(m.group(1)) if m else 0

        for r in row_to_rids:
            row_to_rids[r].sort(key=rid_num)

        total_to_extract = sum(len(rids) for rids in row_to_rids.values())
        log(f"[plan] unique (row, rId) pairs to extract: {total_to_extract}")

        # 5) Extract & resize images one at a time.
        mappings: List[dict] = []
        hash_to_rids: Dict[str, List[str]] = defaultdict(list)
        rid_to_hash: Dict[str, str] = {}
        rid_to_media: Dict[str, str] = {}
        total_out_bytes = 0
        files_written = 0
        skipped_io_errors = 0
        skipped_decode_errors = 0

        # Pre-resolve media paths once per rId
        for rid, target in rel_map.items():
            rid_to_media[rid] = resolve_media_path(target)

        # Iterate rows in ascending order for predictable output
        for row_0idx in sorted(row_to_rids.keys()):
            rids = row_to_rids[row_0idx]
            row_1idx = row_0idx + 1
            raw_model = model_by_row.get(row_1idx)
            clean_model = sanitize_model_no(raw_model)

            for idx, rid in enumerate(rids):
                # First image on a row has no suffix; subsequent get _2, _3...
                suffix = "" if idx == 0 else f"_{idx + 1}"
                media_archive_path = rid_to_media[rid]
                src_ext = file_extension_for(media_archive_path)

                try:
                    raw_bytes = zf.read(media_archive_path)
                except KeyError:
                    log(f"  ! missing media entry for {rid} "
                        f"-> {media_archive_path} (row {row_1idx})")
                    skipped_io_errors += 1
                    continue

                # Source-image hash (over the raw archive bytes, before resize)
                src_hash = hashlib.sha256(raw_bytes).hexdigest()
                if rid not in rid_to_hash:
                    rid_to_hash[rid] = src_hash
                    hash_to_rids[src_hash].append(rid)

                # Resize / re-encode
                try:
                    out_bytes, out_ext, src_w, src_h, out_w, out_h = (
                        resize_image_bytes(raw_bytes, src_ext)
                    )
                except Exception as e:
                    log(f"  ! decode error for {rid} (row {row_1idx}): {e}")
                    skipped_decode_errors += 1
                    continue

                filename = f"{clean_model}_{row_1idx}{suffix}.{out_ext}"
                out_path = os.path.join(OUTPUT_DIR, filename)

                # Avoid filename collisions (extremely unlikely but safe)
                if os.path.exists(out_path):
                    base, ext = os.path.splitext(filename)
                    cand = f"{base}__{src_hash[:8]}{ext}"
                    out_path = os.path.join(OUTPUT_DIR, cand)
                    filename = cand

                try:
                    with open(out_path, "wb") as out_f:
                        out_f.write(out_bytes)
                except OSError as e:
                    log(f"  ! write error for {out_path}: {e}")
                    skipped_io_errors += 1
                    continue

                total_out_bytes += len(out_bytes)
                files_written += 1
                mappings.append(
                    {
                        "row": row_1idx,
                        "model_no": raw_model if raw_model else "",
                        "model_no_clean": clean_model,
                        "image_file": filename,
                        "image_path": out_path,
                        "source_rid": rid,
                        "source_media_path": media_archive_path,
                        "source_sha256": src_hash,
                        "source_size_bytes": len(raw_bytes),
                        "output_size_bytes": len(out_bytes),
                        "source_dimensions": [src_w, src_h],
                        "output_dimensions": [out_w, out_h],
                        "output_ext": out_ext,
                        "resized": max(src_w, src_h) > MAX_LONG_EDGE,
                        "multi_image_row": len(rids) > 1,
                        "image_index_in_row": idx + 1,
                    }
                )

                if files_written % 50 == 0:
                    log(f"  ... wrote {files_written}/{total_to_extract} files")

        # 6) Identify "linked" duplicates: rows that share the same source
        #    image bytes (different rIds but identical bytes), as well as rows
        #    that share the same rId.
        # Group mapping indices by source_sha256 to populate shared-source links.
        hash_to_mapping_idxs: Dict[str, List[int]] = defaultdict(list)
        for mi, mv in enumerate(mappings):
            hash_to_mapping_idxs[mv["source_sha256"]].append(mi)

        for _h, idxs in hash_to_mapping_idxs.items():
            rows_shared = sorted({mappings[mi]["row"] for mi in idxs})
            is_dup = len(idxs) > 1
            for mi in idxs:
                mappings[mi]["shared_source_rows"] = rows_shared
                mappings[mi]["is_duplicate_source"] = is_dup

        unique_source_hashes = {m["source_sha256"] for m in mappings}
        unique_rids = {m["source_rid"] for m in mappings}

        avg_out_bytes = (
            total_out_bytes / files_written if files_written else 0.0
        )

        manifest = {
            "source_xlsx": XLSX_PATH,
            "sheet_name": SHEET_NAME,
            "product_row_range": [PRODUCT_FIRST_ROW, PRODUCT_LAST_ROW],
            "total_product_rows": PRODUCT_LAST_ROW - PRODUCT_FIRST_ROW + 1,
            "product_rows_with_images": len(row_to_rids),
            "product_rows_without_images": (
                PRODUCT_LAST_ROW - PRODUCT_FIRST_ROW + 1 - len(row_to_rids)
            ),
            "total_images_extracted": files_written,
            "total_unique_images": len(unique_source_hashes),
            "total_unique_source_rids": len(unique_rids),
            "total_media_files_in_xlsx": len(media_entries),
            "media_ext_counts": dict(ext_counts),
            "output_dir": OUTPUT_DIR,
            "max_long_edge_px": MAX_LONG_EDGE,
            "average_output_size_bytes": round(avg_out_bytes, 1),
            "average_output_size_kb": round(avg_out_bytes / 1024.0, 1),
            "total_output_size_bytes": total_out_bytes,
            "total_output_size_mb": round(
                total_out_bytes / 1024.0 / 1024.0, 2
            ),
            "issues": {
                "anchors_without_rid": skipped_no_rid,
                "anchors_with_broken_rid": skipped_broken_rid,
                "anchors_in_non_image_col": skipped_wrong_col,
                "anchors_outside_product_range": skipped_out_of_range,
                "media_io_errors": skipped_io_errors,
                "image_decode_errors": skipped_decode_errors,
                "multi_image_rows": sum(
                    1 for rids in row_to_rids.values() if len(rids) > 1
                ),
                "duplicate_source_groups": sum(
                    1 for _h, idxs in hash_to_mapping_idxs.items()
                    if len(idxs) > 1
                ),
            },
            "mappings": mappings,
        }

        with open(MANIFEST_PATH, "w", encoding="utf-8") as mf:
            json.dump(manifest, mf, indent=2, ensure_ascii=False)

        log("=" * 70)
        log(f"DONE. Wrote {files_written} image files to {OUTPUT_DIR}")
        log(f"     Unique source images (by sha256): "
            f"{len(unique_source_hashes)}")
        log(f"     Unique source rIds used         : {len(unique_rids)}")
        log(f"     Total output size               : "
            f"{total_out_bytes / 1024 / 1024:.2f} MB")
        log(f"     Average output size             : "
            f"{avg_out_bytes / 1024:.1f} KB")
        log(f"     Manifest written                : {MANIFEST_PATH}")
        log("=" * 70)
        return 0


if __name__ == "__main__":
    sys.exit(main())
