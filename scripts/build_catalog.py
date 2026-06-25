#!/usr/bin/env python3
"""
Build a unified product catalog by merging:
1. PDF products (with tier classification) — source of truth for catalog
2. HTML inventory (with stock quantities) — source of truth for stock
3. XLSX images manifest — source of truth for images
4. XLSX pricing (MRP, SSP, discount) — enriched pricing

Output: /home/z/my-project/src/data/catalog.json (the master product list)
"""
import json
import re
import os
import shutil
from pathlib import Path
from collections import defaultdict

# ---------------- Paths ----------------
ROOT = Path("/home/z/my-project")
PDF_ANALYSIS = ROOT / "scripts/output/ssp_pdf_analysis.json"
IMAGE_MANIFEST = ROOT / "scripts/output/image_manifest.json"
HTML_INVENTORY = ROOT / "upload/LaxRee_Inventory_2026-06-18 (1) (3).html"
XLSX_FILE = ROOT / "upload/SSP Final OLD (12).xlsx"

CATALOG_OUT = ROOT / "src/data/catalog.json"
PUBLIC_IMAGES = ROOT / "public/product-images"
PUBLIC_IMAGES.mkdir(parents=True, exist_ok=True)

# ---------------- Helpers ----------------
def normalize_model(s: str) -> str:
    """Normalize a model code: LRMB - 132 -> LRMB-132, lowercase, collapse spaces."""
    if not s:
        return ""
    s = str(s).upper().strip()
    # Extract LR prefix + digits
    m = re.search(r'(LR[A-Z]{2})\s*-?\s*(\d+)', s)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    # Fallback: collapse whitespace and remove parens
    s = re.sub(r'\s+', '-', s.strip())
    s = re.sub(r'[^A-Z0-9\-]', '', s)
    return s

def normalize_color(s: str) -> str:
    if not s:
        return ""
    s = str(s).strip()
    placeholders = {"AS PER PICTURE", "AS PER SELECTION", "AS PER CATALOGUE", "AS PER CATALOG"}
    if s.upper() in placeholders:
        return "Multi"
    return s

def parse_size_from_description(desc: str) -> str:
    """Extract size string from description."""
    if not desc:
        return ""
    # Look for size patterns like "290 X 165 X 215 MM" or "13.5*16.3*7cm"
    m = re.search(r'(\d+(?:\.\d+)?\s*[*xX×]\s*\d+(?:\.\d+)?(?:\s*[*xX×]\s*\d+(?:\.\d+)?)?\s*(?:mm|cm|m|MM|CM|M)?)', desc)
    if m:
        return m.group(1).strip()
    # Also check for "Size:" prefix
    m = re.search(r'size[:\s]+([^.)]+)', desc, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    return ""

# ---------------- 1. Load PDF products ----------------
print("1. Loading PDF products...")
with open(PDF_ANALYSIS) as f:
    pdf_data = json.load(f)

pdf_products = pdf_data["all_products"]
print(f"   Loaded {len(pdf_products)} PDF products")

# ---------------- 2. Load HTML inventory stock ----------------
print("2. Loading HTML inventory stock...")
# Read HTML file and extract EMBEDDED_DATA JSON
html_text = HTML_INVENTORY.read_text(encoding='utf-8', errors='ignore')

# Find the EMBEDDED_DATA = {...}; block
m = re.search(r'EMBEDDED_DATA\s*=\s*(\{.*?\});\s*\n', html_text, re.DOTALL)
if not m:
    raise RuntimeError("Could not find EMBEDDED_DATA in HTML")

embedded_data = json.loads(m.group(1))
smap = embedded_data.get("smap", {})
print(f"   Loaded {len(smap)} inventory records")

# Build stock lookup by normalized model code (and by model+color)
stock_by_model = defaultdict(list)  # model_no -> [stock_records]
for key, rec in smap.items():
    model_norm = normalize_model(rec.get("model", ""))
    rec["_key"] = key
    rec["_model_norm"] = model_norm
    rec["_color_norm"] = normalize_color(rec.get("colour", ""))
    stock_by_model[model_norm].append(rec)

# ---------------- 3. Load image manifest ----------------
print("3. Loading image manifest...")
with open(IMAGE_MANIFEST) as f:
    img_data = json.load(f)

# Build image lookup by normalized model_no
images_by_model = defaultdict(list)
for m in img_data["mappings"]:
    model_clean = m.get("model_no_clean", "")
    model_norm = normalize_model(model_clean.replace("_", "-"))
    if model_norm:
        images_by_model[model_norm].append(m)

print(f"   Loaded {len(img_data['mappings'])} image mappings, {len(images_by_model)} unique models")

# ---------------- 4. Load XLSX pricing data (SSP, MRP, Color, Description) ----------------
print("4. Loading XLSX pricing data...")
# Lazy import
try:
    from openpyxl import load_workbook
    wb = load_workbook(XLSX_FILE, read_only=True, keep_links=False, data_only=True)
    ws = wb["New Laxreee data"]
    
    pricing_by_model = defaultdict(list)
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, max_row=900, values_only=True), start=4):
        if not row or len(row) < 7:
            continue
        model_no = str(row[0] or "").strip()
        if not model_no:
            continue
        model_norm = normalize_model(model_no)
        pricing_by_model[model_norm].append({
            "row": row_idx,
            "model_no_raw": model_no,
            "description": str(row[1] or "").strip(),
            "discount_abs": row[2],
            "discount_pct": row[3],
            "mrp": row[4],
            "ssp": row[5],
            "color_raw": str(row[6] or "").strip(),
        })
    wb.close()
    print(f"   Loaded pricing for {len(pricing_by_model)} models")
except Exception as e:
    print(f"   WARNING: Could not load XLSX pricing — {e}")
    pricing_by_model = {}

# ---------------- 5. Build unified catalog ----------------
print("5. Building unified catalog...")
catalog = []
seen_models = set()

# 5a. PDF products (with tiers) — primary catalog
for p in pdf_products:
    model_norm = p.get("model_no_norm") or normalize_model(p.get("model_no", ""))
    if not model_norm:
        continue
    
    # Find stock
    stock_records = stock_by_model.get(model_norm, [])
    if stock_records:
        # Sum balance across all color variants
        total_balance = sum(r.get("balance", 0) for r in stock_records)
        # Pick the best-matching stock record by color
        color = normalize_color(p.get("color", ""))
        best_stock = None
        for sr in stock_records:
            if color and sr.get("_color_norm", "").upper() == color.upper():
                best_stock = sr
                break
        if best_stock is None:
            best_stock = stock_records[0]
        
        stock_qty = best_stock.get("balance", 0)
        inward = best_stock.get("inward", 0)
        dispatched = best_stock.get("dispatched", 0)
    else:
        stock_qty = 0
        inward = 0
        dispatched = 0
        total_balance = 0
    
    # Find image
    image_match = None
    image_candidates = images_by_model.get(model_norm, [])
    if image_candidates:
        # Prefer non-duplicate, smallest image_index_in_row
        image_candidates.sort(key=lambda x: (x.get("image_index_in_row", 99), x.get("row", 9999)))
        image_match = image_candidates[0]
    
    # Find pricing (from XLSX for richer data — discount, MRP)
    pricing_match = None
    pricing_candidates = pricing_by_model.get(model_norm, [])
    if pricing_candidates:
        # Pick first one with valid MRP/SSP
        for pc in pricing_candidates:
            if pc.get("mrp") or pc.get("ssp"):
                pricing_match = pc
                break
        if not pricing_match:
            pricing_match = pricing_candidates[0]
    
    # Determine image path
    image_path = None
    if image_match:
        src_path = Path(image_match["image_path"])
        if src_path.exists():
            # Copy/symlink to public folder
            dest_name = src_path.name
            dest_path = PUBLIC_IMAGES / dest_name
            if not dest_path.exists():
                try:
                    shutil.copy2(src_path, dest_path)
                except Exception as e:
                    print(f"   WARN: copy failed for {src_path}: {e}")
            image_path = f"/product-images/{dest_name}"
    
    # Build product record
    ssp = p.get("ssp")
    if ssp is None and pricing_match:
        ssp = pricing_match.get("ssp")
    if ssp is not None:
        try:
            ssp = round(float(ssp), 2)
        except (ValueError, TypeError):
            ssp = None
    
    mrp = None
    if pricing_match and pricing_match.get("mrp"):
        try:
            mrp = round(float(pricing_match["mrp"]), 2)
        except (ValueError, TypeError):
            pass
    
    discount_pct = None
    if pricing_match and pricing_match.get("discount_pct"):
        try:
            discount_pct = round(float(pricing_match["discount_pct"]), 2)
        except (ValueError, TypeError):
            pass
    
    description = p.get("description", "") or (pricing_match.get("description", "") if pricing_match else "")
    
    product = {
        "id": f"{model_norm}-{p.get('page', 0)}-{len(catalog)+1}",
        "model_no": model_norm,
        "model_no_raw": p.get("model_no", ""),
        "name": p.get("product", "") or (pricing_match.get("description", "").split('\n')[0] if pricing_match else ""),
        "category": p.get("product_filled") or p.get("product", ""),
        "tier": p.get("tier_filled") or p.get("tier", "Standard"),
        "description": description,
        "color": normalize_color(p.get("color", "")),
        "size": parse_size_from_description(description),
        "ssp": ssp,
        "mrp": mrp,
        "discount_pct": discount_pct,
        "stock_qty": stock_qty,
        "inward": inward,
        "dispatched": dispatched,
        "image_url": image_path,
        "in_stock": stock_qty > 0,
        "source": "pdf",
    }
    catalog.append(product)
    seen_models.add(model_norm)

# 5b. HTML inventory products NOT in PDF — add as "General" tier
# (So we don't lose stock visibility for items that exist in inventory but not the tiered catalog)
for model_norm, recs in stock_by_model.items():
    if model_norm in seen_models:
        continue
    
    # Pick the first record (or sum)
    total_balance = sum(r.get("balance", 0) for r in recs)
    first_rec = recs[0]
    
    # Find image
    image_match = None
    image_candidates = images_by_model.get(model_norm, [])
    if image_candidates:
        image_candidates.sort(key=lambda x: (x.get("image_index_in_row", 99), x.get("row", 9999)))
        image_match = image_candidates[0]
    
    image_path = None
    if image_match:
        src_path = Path(image_match["image_path"])
        if src_path.exists():
            dest_name = src_path.name
            dest_path = PUBLIC_IMAGES / dest_name
            if not dest_path.exists():
                try:
                    shutil.copy2(src_path, dest_path)
                except Exception:
                    pass
            image_path = f"/product-images/{dest_name}"
    
    # Find pricing
    pricing_match = None
    pricing_candidates = pricing_by_model.get(model_norm, [])
    if pricing_candidates:
        for pc in pricing_candidates:
            if pc.get("mrp") or pc.get("ssp"):
                pricing_match = pc
                break
        if not pricing_match:
            pricing_match = pricing_candidates[0]
    
    ssp = None
    mrp = None
    discount_pct = None
    description = first_rec.get("item", "")
    if pricing_match:
        try:
            ssp = round(float(pricing_match["ssp"]), 2) if pricing_match.get("ssp") else None
        except (ValueError, TypeError):
            pass
        try:
            mrp = round(float(pricing_match["mrp"]), 2) if pricing_match.get("mrp") else None
        except (ValueError, TypeError):
            pass
        try:
            discount_pct = round(float(pricing_match["discount_pct"]), 2) if pricing_match.get("discount_pct") else None
        except (ValueError, TypeError):
            pass
        if pricing_match.get("description"):
            description = pricing_match["description"]
    
    product = {
        "id": f"{model_norm}-INV-{len(catalog)+1}",
        "model_no": model_norm,
        "model_no_raw": first_rec.get("model", ""),
        "name": first_rec.get("item", model_norm),
        "category": first_rec.get("category", "General Inventory"),
        "tier": "Standard",
        "description": description,
        "color": normalize_color(first_rec.get("colour", "")),
        "size": parse_size_from_description(description),
        "ssp": ssp,
        "mrp": mrp,
        "discount_pct": discount_pct,
        "stock_qty": total_balance,
        "inward": sum(r.get("inward", 0) for r in recs),
        "dispatched": sum(r.get("dispatched", 0) for r in recs),
        "image_url": image_path,
        "in_stock": total_balance > 0,
        "source": "inventory",
    }
    catalog.append(product)
    seen_models.add(model_norm)

# 5c. XLSX-only products (no PDF, no inventory stock) — add as "Catalog" tier with 0 stock
for model_norm, pricings in pricing_by_model.items():
    if model_norm in seen_models:
        continue
    
    # Pick the first pricing
    pm = pricings[0]
    
    # Find image
    image_match = None
    image_candidates = images_by_model.get(model_norm, [])
    if image_candidates:
        image_candidates.sort(key=lambda x: (x.get("image_index_in_row", 99), x.get("row", 9999)))
        image_match = image_candidates[0]
    
    image_path = None
    if image_match:
        src_path = Path(image_match["image_path"])
        if src_path.exists():
            dest_name = src_path.name
            dest_path = PUBLIC_IMAGES / dest_name
            if not dest_path.exists():
                try:
                    shutil.copy2(src_path, dest_path)
                except Exception:
                    pass
            image_path = f"/product-images/{dest_name}"
    
    try:
        ssp = round(float(pm.get("ssp")), 2) if pm.get("ssp") else None
    except (ValueError, TypeError):
        ssp = None
    try:
        mrp = round(float(pm.get("mrp")), 2) if pm.get("mrp") else None
    except (ValueError, TypeError):
        mrp = None
    try:
        discount_pct = round(float(pm.get("discount_pct")), 2) if pm.get("discount_pct") else None
    except (ValueError, TypeError):
        discount_pct = None
    
    description = pm.get("description", "")
    product = {
        "id": f"{model_norm}-XLSX-{len(catalog)+1}",
        "model_no": model_norm,
        "model_no_raw": pm.get("model_no_raw", ""),
        "name": description.split('\n')[0][:80] if description else model_norm,
        "category": "Catalog Item",
        "tier": "Standard",
        "description": description,
        "color": normalize_color(pm.get("color_raw", "")),
        "size": parse_size_from_description(description),
        "ssp": ssp,
        "mrp": mrp,
        "discount_pct": discount_pct,
        "stock_qty": 0,
        "inward": 0,
        "dispatched": 0,
        "image_url": image_path,
        "in_stock": False,
        "source": "xlsx",
    }
    catalog.append(product)
    seen_models.add(model_norm)

# ---------------- 6. Stats ----------------
print("\n6. Final catalog stats:")
print(f"   Total products: {len(catalog)}")
by_tier = defaultdict(int)
by_source = defaultdict(int)
by_category = defaultdict(int)
with_image = 0
with_stock = 0
with_price = 0
for p in catalog:
    by_tier[p["tier"]] += 1
    by_source[p["source"]] += 1
    by_category[p["category"]] += 1
    if p["image_url"]:
        with_image += 1
    if p["stock_qty"] > 0:
        with_stock += 1
    if p["ssp"]:
        with_price += 1

print(f"   By tier: {dict(by_tier)}")
print(f"   By source: {dict(by_source)}")
print(f"   With image: {with_image} ({100*with_image/len(catalog):.1f}%)")
print(f"   With stock > 0: {with_stock} ({100*with_stock/len(catalog):.1f}%)")
print(f"   With price: {with_price} ({100*with_price/len(catalog):.1f}%)")
print(f"   Categories ({len(by_category)}): {dict(sorted(by_category.items(), key=lambda x: -x[1]))}")

# ---------------- 7. Save catalog ----------------
CATALOG_OUT.parent.mkdir(parents=True, exist_ok=True)
with open(CATALOG_OUT, 'w', encoding='utf-8') as f:
    json.dump(catalog, f, ensure_ascii=False, indent=2)
print(f"\n7. Catalog saved to: {CATALOG_OUT}")
print(f"   Size: {CATALOG_OUT.stat().st_size / 1024:.1f} KB")

# Also save categories summary
summary = {
    "total_products": len(catalog),
    "by_tier": dict(by_tier),
    "by_source": dict(by_source),
    "by_category": dict(by_category),
    "with_image": with_image,
    "with_stock": with_stock,
    "with_price": with_price,
    "categories_list": sorted(by_category.keys()),
    "tiers_list": sorted([t for t in by_tier.keys() if t]),
}
with open(ROOT / "src/data/catalog_summary.json", 'w', encoding='utf-8') as f:
    json.dump(summary, f, ensure_ascii=False, indent=2)
print(f"   Summary: {ROOT / 'src/data/catalog_summary.json'}")

print("\nDONE.")
